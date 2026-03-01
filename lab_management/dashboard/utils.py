"""
Utility functions for lab management dashboard.

This module contains helper functions for Excel import and data processing.
"""


import openpyxl
from django.db import transaction
from .models import Cluster, Server


def import_from_excel(file_path):
    """
    Import clusters and nodes from Excel file.
    Expected columns for clusters: CLUSTER NAME, OWNER, DESCRIPTION, GPU COUNT, GPU TYPE, IB BAND
    Expected columns for nodes: ROLE, SERVER MODEL, GENERATION, SERVICE TAG, IDRAC IP,
    IDRAC CREDS, BMC MAC ADDRESS, PXE MAC ADDRESS, CLUSTER NAME
    """
    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # Find header row
        header_row = None
        for row in sheet.iter_rows():
            if any(cell.value and ('ROLE' in str(cell.value).upper() or 'SERVER_MODEL' in str(cell.value).upper()) for cell in row):
                header_row = row
                break

        if not header_row:
            raise ValueError("Header row not found. Expected 'ROLE' or 'SERVER_MODEL' column.")

        # Get column indices
        col_indices = {}
        for i, cell in enumerate(header_row):
            if cell.value:
                col_name = str(cell.value).strip().upper()
                # Cluster columns
                if 'CLUSTER NAME' in col_name:
                    col_indices['cluster_name'] = i
                elif 'OWNER' in col_name and 'CLUSTER' not in col_name:
                    col_indices['owner'] = i
                elif 'DESCRIPTION' in col_name and 'CLUSTER' not in col_name:
                    col_indices['description'] = i
                elif 'GPU COUNT' in col_name:
                    col_indices['gpu_count'] = i
                elif 'GPU TYPE' in col_name:
                    col_indices['gpu_type'] = i
                elif 'IB BAND' in col_name or 'INFINIBAND' in col_name:
                    col_indices['ib_band'] = i
                # Server columns (matching your Excel format)
                elif 'ROLE' in col_name:
                    col_indices['role'] = i
                elif 'SERVER_MODEL' in col_name or 'SERVER MODEL' in col_name:
                    col_indices['server_model'] = i
                elif 'GENERATION' in col_name:
                    col_indices['generation'] = i
                elif 'SERVICE TAG' in col_name or 'SERVICE_TAG' in col_name:
                    col_indices['service_tag'] = i
                elif 'IDRAC IP' in col_name or 'IDRAC_IP' in col_name:
                    col_indices['idrac_ip'] = i
                elif 'IDRAC CREDS' in col_name or 'IDRAC_CREDS' in col_name:
                    col_indices['idrac_creds'] = i
                elif 'BMC MAC ADDRESS' in col_name or 'BMC_MAC_ADDRESS' in col_name:
                    col_indices['bmc_mac'] = i
                elif 'PXE MAC ADDRESS' in col_name or 'PXE_MAC_ADDRESS' in col_name:
                    col_indices['pxe_mac'] = i

        imported_clusters = 0
        imported_nodes = 0
        errors = []
        clusters_cache = {}

        with transaction.atomic():
            for row in sheet.iter_rows(min_row=header_row[0].row + 1):
                if not any(cell.value for cell in row):
                    continue

                try:
                    # Check if this is a cluster row (has cluster name but no service tag)
                    cluster_name = str(row[col_indices.get('cluster_name', 0)].value or '').strip() if col_indices.get('cluster_name') is not None else ''
                    service_tag = str(row[col_indices.get('service_tag', 3)].value or '').strip() if col_indices.get('service_tag') is not None else ''

                    # If cluster name exists but no service tag, treat as cluster definition
                    if cluster_name and not service_tag:
                        owner = str(row[col_indices.get('owner', 1)].value or '').strip() if col_indices.get('owner') is not None else ''
                        description = str(row[col_indices.get('description', 2)].value or '').strip() if col_indices.get('description') is not None else ''
                        gpu_count = str(row[col_indices.get('gpu_count', 3)].value or '').strip() if col_indices.get('gpu_count') is not None else '0'
                        gpu_type = str(row[col_indices.get('gpu_type', 4)].value or '').strip() if col_indices.get('gpu_type') is not None else ''
                        ib_band = str(row[col_indices.get('ib_band', 5)].value or '').strip() if col_indices.get('ib_band') is not None else ''

                        # Create or update cluster
                        cluster, created = Cluster.objects.update_or_create(
                            name=cluster_name,
                            defaults={
                                'owner': owner if owner else 'Unknown',
                                'description': description,
                                'gpu_count': int(gpu_count) if gpu_count.isdigit() else 0,
                                'gpu_type': gpu_type,
                                'ib_band': ib_band
                            }
                        )

                        if created:
                            imported_clusters += 1
                        clusters_cache[cluster_name] = cluster
                        continue

                    # If service tag exists, treat as server row
                    if service_tag:
                        # Extract server data
                        role = str(row[col_indices.get('role', 0)].value or '').strip() if col_indices.get('role') is not None else ''
                        server_model = str(row[col_indices.get('server_model', 1)].value or '').strip() if col_indices.get('server_model') is not None else ''
                        generation = str(row[col_indices.get('generation', 2)].value or '').strip() if col_indices.get('generation') is not None else ''
                        idrac_ip = str(row[col_indices.get('idrac_ip', 4)].value or '').strip() if col_indices.get('idrac_ip') is not None else ''
                        idrac_creds = str(row[col_indices.get('idrac_creds', 5)].value or '').strip() if col_indices.get('idrac_creds') is not None else ''
                        bmc_mac = str(row[col_indices.get('bmc_mac', 6)].value or '').strip() if col_indices.get('bmc_mac') is not None else ''
                        pxe_mac = str(row[col_indices.get('pxe_mac', 7)].value or '').strip() if col_indices.get('pxe_mac') is not None else ''
                        server_cluster_name = str(row[col_indices.get('cluster_name', 8)].value or '').strip() if col_indices.get('cluster_name') is not None else ''

                        # Determine cluster
                        if server_cluster_name and server_cluster_name in clusters_cache:
                            cluster = clusters_cache[server_cluster_name]
                        elif server_cluster_name:
                            # Try to find existing cluster or create new one
                            cluster, created = Cluster.objects.get_or_create(
                                name=server_cluster_name,
                                defaults={
                                    'owner': 'Unknown',
                                    'description': 'Auto-created from server import'
                                }
                            )
                            if created:
                                imported_clusters += 1
                            clusters_cache[server_cluster_name] = cluster
                        else:
                            # Use first available cluster or create default
                            if clusters_cache:
                                cluster = list(clusters_cache.values())[0]
                            else:
                                cluster, created = Cluster.objects.get_or_create(
                                    name='Default Cluster',
                                    defaults={
                                        'owner': 'System Admin',
                                        'description': 'Default cluster for imported nodes'
                                    }
                                )
                                if created:
                                    imported_clusters += 1
                                clusters_cache['Default Cluster'] = cluster

                        # Check if node with this service tag already exists in any cluster
                        existing_node = Server.objects.filter(service_tag__iexact=service_tag).first()
                        if existing_node:
                            errors.append(f"Row {row[0].row}: Node with service tag '{service_tag}' already exists in cluster '{existing_node.cluster.name}'. Delete it from there first before importing.")
                        else:
                            # Create new node
                            node = Server.objects.create(
                                cluster=cluster,
                                role=role.lower() if role else 'worker',
                                server_model=server_model,
                                generation=generation,
                                service_tag=service_tag,
                                idrac_ip=idrac_ip if idrac_ip else None,
                                idrac_creds=idrac_creds,
                                bmc_mac_address=bmc_mac,
                                pxe_mac_address=pxe_mac,
                            )
                            imported_nodes += 1
                    else:
                        errors.append(f"Row {row[0].row}: Neither cluster name nor service tag found")

                except Exception as e:
                    errors.append(f"Row {row[0].row}: {str(e)}")

        return {
            'success': True,
            'imported_clusters': imported_clusters,
            'imported_nodes': imported_nodes,
            'errors': errors
        }

    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }
